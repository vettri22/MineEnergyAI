"""
reports/excel_generator.py
-----------------------------
Generates a professionally formatted Excel (.xlsx) report using openpyxl,
including a summary sheet and a detailed per-machine data sheet.
"""

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import Config

HEADER_FILL = PatternFill(start_color="0B5FFF", end_color="0B5FFF", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=16, color="101828")
SUBTITLE_FONT = Font(size=10, color="667085")
THIN_BORDER = Border(*(Side(style="thin", color="E0E4EA"),) * 4)


def generate_excel_report(report_type, period_label, kpis, rows, output_dir=None):
    """
    Build an Excel workbook report and return its full path.
    Sheet 1: Executive Summary (KPIs)
    Sheet 2: Machine-wise energy data
    """
    output_dir = output_dir or Config.REPORTS_FOLDER
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"MineEnergyAI_{report_type}_Report_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()

    # ---------- Sheet 1: Summary ----------
    ws = wb.active
    ws.title = "Executive Summary"
    ws.merge_cells("A1:D1")
    ws["A1"] = "MineEnergy AI — Intelligent Energy Management"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:D2")
    ws["A2"] = f"{report_type} Report | Period: {period_label}"
    ws["A2"].font = SUBTITLE_FONT

    kpi_rows = [
        ("Total Energy (kWh)", kpis["total_energy"]),
        ("Electricity Cost (INR)", kpis["total_cost"]),
        ("Carbon Emission (kg CO2)", kpis["total_carbon"]),
        ("Average Efficiency Score", kpis["avg_efficiency"]),
        ("Machines Monitored", kpis["machine_count"]),
        ("Open Alerts", kpis["alert_count"]),
    ]
    start_row = 4
    ws.cell(row=start_row, column=1, value="Metric").font = HEADER_FONT
    ws.cell(row=start_row, column=1).fill = HEADER_FILL
    ws.cell(row=start_row, column=2, value="Value").font = HEADER_FONT
    ws.cell(row=start_row, column=2).fill = HEADER_FILL

    for i, (label, value) in enumerate(kpi_rows, start=start_row + 1):
        ws.cell(row=i, column=1, value=label).border = THIN_BORDER
        ws.cell(row=i, column=2, value=value).border = THIN_BORDER

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    # ---------- Sheet 2: Machine data ----------
    ws2 = wb.create_sheet("Machine Data")
    headers = ["Machine Code", "Machine Name", "Type", "Energy (kWh)",
               "Ore Processed (T)", "Efficiency Score", "Status"]
    for col, header in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for r_idx, row in enumerate(rows, start=2):
        values = [
            row["machine_code"], row["machine_name"], row["machine_type"],
            row["energy_kwh"], row["ore_processed"], row["efficiency_score"], row["status"],
        ]
        for c_idx, value in enumerate(values, start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            cell.border = THIN_BORDER
            if r_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F4F6FA", end_color="F4F6FA", fill_type="solid")

    for col in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 20

    ws2.freeze_panes = "A2"

    wb.save(filepath)
    return filepath
